from django.shortcuts import render, redirect, HttpResponse
from django.urls import reverse_lazy
from django.views.generic import (
    ListView,
    DetailView,
    FormView,
    View,
    CreateView,
    UpdateView,
)
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Q
from django.db import connection
from django.db import transaction
from django.utils import timezone
from django.contrib.auth.mixins import UserPassesTestMixin
from django.utils.crypto import get_random_string
from django.core.exceptions import PermissionDenied
import openpyxl
from io import BytesIO

from .models import Employee
from .forms import EmployeeImportForm, EmployeeForm, HROnboardingForm
from apps.organization.models import Organization
from utils.sql_scope import (
    build_org_tree_cte,
    get_user_scope,
    normalize_str,
    parse_iso_date,
    uniq,
)


# 1. 导入处理 View
class EmployeeImportView(LoginRequiredMixin, FormView):
    template_name = "employee/import.html"
    form_class = EmployeeImportForm
    success_url = "/employee/"

    def form_valid(self, form):
        file = form.cleaned_data["file"]
        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            success_count = 0
            errors = []

            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if not row[0]:
                    # 忽略空行
                    continue
                try:
                    (
                        name,
                        emp_id,
                        id_card,
                        birth_date,
                        phone,
                        email,
                        org_code,
                        position,
                        hire_date,
                        gender,
                    ) = row[:10]
                except ValueError:
                    errors.append(f"第 {row_idx} 行格式错误：列数不足")
                    continue

                if Employee.objects.filter(emp_id=emp_id).exists():
                    errors.append(f"第 {row_idx} 行错误：工号 {emp_id} 已存在")
                    continue
                if not birth_date:
                    errors.append(f"第 {row_idx} 行错误：出生日期不能为空")
                    continue
                if not hire_date:
                    errors.append(f"第 {row_idx} 行错误：入职日期不能为空")
                    continue

                org = Organization.objects.filter(org_code=org_code).first()
                if not org:
                    errors.append(f"第 {row_idx} 行错误：部门编码 {org_code} 不存在")
                    continue

                try:
                    with transaction.atomic():
                        user, _ = User.objects.get_or_create(
                            username=emp_id, defaults={"email": email}
                        )
                        user.set_password("123456")
                        user.save()

                        if isinstance(birth_date, str):
                            birth_date = timezone.datetime.strptime(
                                birth_date, "%Y-%m-%d"
                            ).date()
                        if isinstance(hire_date, str):
                            hire_date = timezone.datetime.strptime(
                                hire_date, "%Y-%m-%d"
                            ).date()

                        Employee.objects.create(
                            emp_id=str(emp_id),
                            emp_name=name,
                            id_card=str(id_card),
                            birth_date=birth_date,
                            phone=str(phone),
                            email=email,
                            org=org,
                            position=position,
                            hire_date=hire_date,
                            gender=gender if gender in ["male", "female"] else "male",
                            user=user,
                            employment_type="full_time",
                            emp_status="probation",
                            create_by=self.request.user.username,
                            update_by=self.request.user.username,
                        )
                        success_count += 1
                except Exception as e:
                    errors.append(f"第 {row_idx} 行入库失败：{str(e)}")

            if success_count > 0:
                messages.success(self.request, f"成功导入 {success_count} 名员工！")
            if errors:
                messages.warning(
                    self.request,
                    "部分失败：<br>" + "<br>".join(errors[:5]),
                )

        except Exception as e:
            messages.error(self.request, f"文件解析失败: {str(e)}")
            return self.form_invalid(form)
        return redirect(self.success_url)


# 2. 模板下载 View
class EmployeeTemplateDownloadView(LoginRequiredMixin, View):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "员工导入模板"
        headers = [
            "姓名",
            "工号",
            "身份证号",
            "出生日期",
            "手机号",
            "邮箱",
            "部门编码",
            "岗位",
            "入职日期",
            "性别(male/female)",
        ]
        ws.append(headers)
        ws.append(
            [
                "张三",
                "IMP9001",
                "110101199001011234",
                "1990-01-01",
                "13800000000",
                "zhangsan@test.com",
                "HR",
                "人事专员",
                "2024-01-01",
                "male",
            ]
        )
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=employee_template.xlsx"
        return response


# 3. 创建员工 View
class EmployeeCreateView(LoginRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = "employee/form.html"
    success_url = reverse_lazy("employee:list")

    def form_valid(self, form):
        last_emp = Employee.objects.order_by("-create_time").first()
        new_id_suffix = 1001
        if last_emp and last_emp.emp_id.startswith("EMP"):
            try:
                new_id_suffix = int(last_emp.emp_id[3:]) + 1
            except ValueError:
                pass
        emp_id = f"EMP{new_id_suffix}"

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=emp_id,
                    email=form.cleaned_data.get("email"),
                    password="password123",
                )
                form.instance.emp_id = emp_id
                form.instance.user = user
                form.instance.create_by = self.request.user.username
                form.instance.update_by = self.request.user.username
                messages.success(self.request, f"员工创建成功，工号：{emp_id}")
                return super().form_valid(form)
        except Exception as e:
            messages.error(self.request, f"创建失败: {e}")
            return self.form_invalid(form)


# 4. 更新员工 View
class EmployeeUpdateView(LoginRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = "employee/form.html"

    def get_success_url(self):
        return reverse_lazy("employee:detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        form.instance.update_by = self.request.user.username
        messages.success(self.request, "员工信息已更新")
        return super().form_valid(form)


# 5. 列表 View
class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = "employee/list.html"
    context_object_name = "employees"
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related("org", "manager_emp")
        search = self.request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(emp_name__icontains=search)
                | Q(emp_id__icontains=search)
                | Q(org__org_name__icontains=search)
            )
        return queryset.order_by("emp_id")


class EmployeeOrgSqlSearchView(LoginRequiredMixin, View):
    """原生 SQL：支持上级部门查询下级部门员工，多条件组合；下级不可越权查上级。"""

    template_name = "employee/sql_search.html"
    max_rows = 200

    def get(self, request):
        scope = get_user_scope(
            user_id=request.user.id,
            is_superuser=request.user.is_superuser,
            is_staff=request.user.is_staff,
        )

        if not (scope.is_superuser or scope.is_hr or scope.is_manager):
            raise PermissionDenied

        # Ensure SQL view exists (prevents confusing DB errors)
        with connection.cursor() as cursor:
            cursor.execute("SELECT to_regclass('vw_employee_profile')")
            if cursor.fetchone()[0] is None:
                messages.error(
                    request,
                    "缺少数据库视图 vw_employee_profile：请先运行 hrms/apply_views.py 初始化视图。",
                )
                return redirect("core:dashboard")

        # 允许范围：本人所在组织 + 其负责组织（如有）的子树
        root_ids: list[str] = []
        if scope.org_id:
            root_ids.append(scope.org_id)

        if scope.emp_pk:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT org.id
                    FROM organization org
                    WHERE org.is_deleted = FALSE
                      AND org.manager_emp_id = %s
                    """,
                    [scope.emp_pk],
                )
                root_ids.extend([r[0] for r in cursor.fetchall()])

        root_ids = uniq([rid for rid in root_ids if rid])

        # 组织下拉（仅允许范围内）
        org_options: list[dict[str, str]] = []
        if scope.is_superuser or scope.is_hr:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT id, org_name
                    FROM organization
                    WHERE is_deleted = FALSE
                    ORDER BY org_name
                    """
                )
                org_options = [{"id": r[0], "name": r[1]} for r in cursor.fetchall()]
        else:
            cte_sql, cte_params = build_org_tree_cte(root_ids)
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""
                    {cte_sql}
                    SELECT o.id, o.org_name
                    FROM organization o
                    WHERE o.is_deleted = FALSE
                      AND o.id IN (SELECT id FROM org_tree)
                    ORDER BY o.org_name
                    """,
                    cte_params,
                )
                org_options = [{"id": r[0], "name": r[1]} for r in cursor.fetchall()]

        # Filters
        filter_org = normalize_str(request.GET.get("org"))
        keyword = normalize_str(request.GET.get("keyword"))
        emp_id_exact = normalize_str(request.GET.get("emp_id"))
        email = normalize_str(request.GET.get("email"))
        phone = normalize_str(request.GET.get("phone"))
        id_card = normalize_str(request.GET.get("id_card"))
        gender = normalize_str(request.GET.get("gender"))
        employment_type = normalize_str(request.GET.get("employment_type"))
        manager_keyword = normalize_str(request.GET.get("manager"))
        emp_status = normalize_str(request.GET.get("emp_status"))
        position = normalize_str(request.GET.get("position"))
        hire_start = parse_iso_date(request.GET.get("hire_start"))
        hire_end = parse_iso_date(request.GET.get("hire_end"))
        birth_start = parse_iso_date(request.GET.get("birth_start"))
        birth_end = parse_iso_date(request.GET.get("birth_end"))

        # Query from view (already filters deleted rows)
        clauses: list[str] = ["1=1"]
        params: list[object] = []

        if not (scope.is_superuser or scope.is_hr):
            cte_sql, cte_params = build_org_tree_cte(root_ids)
            clauses.append("v.org_id IN (SELECT id FROM org_tree)")
            params.extend(cte_params)
        else:
            cte_sql = ""

        if filter_org:
            clauses.append("v.org_id = %s")
            params.append(filter_org)

        if emp_id_exact:
            clauses.append("v.emp_id = %s")
            params.append(emp_id_exact)

        if email:
            clauses.append("LOWER(v.email) LIKE %s")
            params.append(f"%{email.lower()}%")

        if phone:
            clauses.append("v.phone LIKE %s")
            params.append(f"%{phone}%")

        if id_card:
            clauses.append("v.id_card LIKE %s")
            params.append(f"%{id_card}%")

        if gender:
            clauses.append("v.gender = %s")
            params.append(gender)

        if employment_type:
            clauses.append("v.employment_type = %s")
            params.append(employment_type)

        if emp_status:
            clauses.append("v.emp_status = %s")
            params.append(emp_status)

        if position:
            clauses.append("LOWER(v.position) LIKE %s")
            params.append(f"%{position.lower()}%")

        if hire_start:
            clauses.append("v.hire_date >= %s")
            params.append(hire_start)

        if hire_end:
            clauses.append("v.hire_date <= %s")
            params.append(hire_end)

        if birth_start:
            clauses.append("v.birth_date >= %s")
            params.append(birth_start)

        if birth_end:
            clauses.append("v.birth_date <= %s")
            params.append(birth_end)

        if keyword:
            kw = f"%{keyword.lower()}%"
            clauses.append(
                "(LOWER(v.emp_name) LIKE %s OR LOWER(v.emp_id) LIKE %s OR LOWER(v.email) LIKE %s)"
            )
            params.extend([kw, kw, kw])

        if manager_keyword:
            kw = f"%{manager_keyword.lower()}%"
            clauses.append(
                "(LOWER(COALESCE(v.manager_emp_name,'')) LIKE %s OR LOWER(COALESCE(v.manager_emp_code,'')) LIKE %s)"
            )
            params.extend([kw, kw])

        where_sql = " AND ".join(clauses) if clauses else "1=1"

        sql = f"""
            {cte_sql}
            SELECT
                v.emp_pk,
                v.emp_id,
                v.emp_name,
                v.gender,
                v.birth_date,
                v.position,
                v.employment_type,
                v.emp_status,
                v.hire_date,
                v.phone,
                v.email,
                v.org_name,
                v.manager_emp_name,
                v.manager_emp_code
            FROM public.vw_employee_profile v
            WHERE {where_sql}
            ORDER BY v.org_name ASC, v.emp_id ASC
            LIMIT {self.max_rows}
        """

        results: list[dict[str, object]] = []
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            description = cursor.description or []
            col_names = [c[0] for c in description]
            for row in cursor.fetchall():
                results.append(dict(zip(col_names, row)))

        ctx = {
            "orgs": org_options,
            "results": results,
            "result_count": len(results),
            "max_rows": self.max_rows,
            "emp_status_choices": Employee.EMP_STATUS_CHOICES,
            "gender_choices": Employee.GENDER_CHOICES,
            "employment_type_choices": Employee.EMPLOYMENT_TYPE_CHOICES,
            "filters": {
                "org": filter_org,
                "keyword": keyword,
                "emp_id": emp_id_exact,
                "email": email,
                "phone": phone,
                "id_card": id_card,
                "gender": gender,
                "employment_type": employment_type,
                "manager": manager_keyword,
                "emp_status": emp_status,
                "position": position,
                "hire_start": hire_start or "",
                "hire_end": hire_end or "",
                "birth_start": birth_start or "",
                "birth_end": birth_end or "",
            },
        }
        return render(request, self.template_name, ctx)


# 6. 详情 View
class EmployeeDetailView(LoginRequiredMixin, DetailView):
    model = Employee
    template_name = "employee/detail.html"
    context_object_name = "emp"


class HROnboardingView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = "employee/hr_onboarding.html"
    form_class = HROnboardingForm
    success_url = reverse_lazy("employee:hr_onboarding")

    def test_func(self):
        return self._is_hr_user()

    def _user_employee(self):
        return getattr(self.request.user, "employee", None)

    def _is_hr_user(self):
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return True
        emp = self._user_employee()
        if not emp:
            return False
        org_name = (emp.org.org_name or "").lower()
        position = (emp.position or "").upper()
        return "人力" in org_name or "HR" in position

    def _is_hr_director(self):
        emp = self._user_employee()
        if not emp:
            return False
        position_raw = emp.position or ""
        position = position_raw.upper()
        return "总监" in position_raw or "DIRECTOR" in position

    def _generate_emp_id(self):
        last_emp = Employee.objects.order_by("-create_time").first()
        suffix = 1001
        if last_emp and last_emp.emp_id.startswith("EMP"):
            try:
                suffix = int(last_emp.emp_id[3:]) + 1
            except ValueError:
                pass
        return f"EMP{suffix}"

    def _context_data(self, form=None):
        form = form or self.form_class()
        pending_onboardings = Employee.objects.filter(emp_status="probation").order_by(
            "-hire_date"
        )[:5]
        recent_onboardings = Employee.objects.order_by("-create_time")[:5]
        last_creds = self.request.session.pop("last_onboarding_creds", None)
        return {
            "form": form,
            "pending_onboardings": pending_onboardings,
            "recent_onboardings": recent_onboardings,
            "is_hr_director": self._is_hr_director(),
            "last_credentials": last_creds,
        }

    def get(self, request):
        form = self.form_class(
            initial={"employment_type": "full_time", "emp_status": "probation"}
        )
        return render(request, self.template_name, self._context_data(form))

    def post(self, request):
        action = request.POST.get("action")
        if action == "approve_onboarding" and self._is_hr_director():
            emp_pk = request.POST.get("employee_pk")
            if emp_pk:
                Employee.objects.filter(pk=emp_pk).update(
                    emp_status="active", update_by=request.user.username
                )
                messages.success(request, "已将该员工状态修改为在职。")
            return redirect(self.success_url)

        form = self.form_class(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, self._context_data(form))

        password = form.cleaned_data.get("initial_password")
        if not password:
            password = get_random_string(10)

        emp_id = self._generate_emp_id()
        emp = form.save(commit=False)
        emp.emp_id = emp_id
        emp.create_by = request.user.username
        emp.update_by = request.user.username
        with transaction.atomic():
            user = User.objects.create_user(
                username=emp_id, email=form.cleaned_data.get("email"), password=password
            )
            emp.user = user
            emp.save()

        self.request.session["last_onboarding_creds"] = {
            "username": emp_id,
            "password": password,
        }
        messages.success(
            request, f"已为 {emp.emp_name} 创建账号 {emp_id}，初始密码已通过提示展示。"
        )
        return redirect(self.success_url)
